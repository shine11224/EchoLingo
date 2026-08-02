import csv, io, json, os, re
from functools import lru_cache
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parents[3]
WORDLISTS_DIR = BASE_DIR / "resources" / "wordlists" / "wordlists"
COMPILED_DIR = WORDLISTS_DIR / "compiled"
PATTERNS_DIR = WORDLISTS_DIR / "patterns"
USER_DIR = WORDLISTS_DIR / "user"

# Ensure directories exist
for d in [COMPILED_DIR, PATTERNS_DIR, USER_DIR]:
    d.mkdir(parents=True, exist_ok=True)

WORD_RE = re.compile(r"^[a-z][a-z'\-]{1,}$")
REQUIRED_PATTERN_FIELDS = ("id", "name", "template", "triggers", "explanation_cn", "example", "ielts_tip")
PATTERN_UPLOAD_SUFFIXES = {".json", ".csv", ".xlsx"}
BNC_COCA_PATH = WORDLISTS_DIR / "bnc_coca_9k.csv"

_IRREGULAR_INFLECTIONS = {
    "be": {"am", "is", "are", "was", "were", "been", "being"},
    "have": {"has", "had", "having"},
    "do": {"does", "did", "done", "doing"},
    "go": {"goes", "went", "gone", "going"},
    "say": {"says", "said", "saying"},
    "make": {"makes", "made", "making"},
    "take": {"takes", "took", "taken", "taking"},
    "come": {"comes", "came", "coming"},
    "see": {"sees", "saw", "seen", "seeing"},
    "get": {"gets", "got", "gotten", "getting"},
    "give": {"gives", "gave", "given", "giving"},
    "know": {"knows", "knew", "known", "knowing"},
    "think": {"thinks", "thought", "thinking"},
    "find": {"finds", "found", "finding"},
    "run": {"runs", "ran", "running"},
    "write": {"writes", "wrote", "written", "writing"},
    "read": {"reads", "reading"},
    "buy": {"buys", "bought", "buying"},
    "bring": {"brings", "brought", "bringing"},
    "teach": {"teaches", "taught", "teaching"},
    "catch": {"catches", "caught", "catching"},
    "child": {"children"}, "person": {"people"},
    "man": {"men"}, "woman": {"women"}, "mouse": {"mice"},
    "goose": {"geese"}, "tooth": {"teeth"}, "foot": {"feet"},
    "good": {"better", "best"}, "bad": {"worse", "worst"},
    "far": {"farther", "farthest", "further", "furthest"},
}

def _clean_word(token: str) -> str | None:
    word = token.strip().lower()
    if WORD_RE.match(word) and len(word) >= 3:
        return word
    return None

def parse_wordlist_content(content: str):
    """Return (words, invalid_tokens) for uploaded plain text / CSV-like wordlists."""
    words = set()
    invalid = []
    for token in re.split(r"[\s,;]+", content):
        raw = token.strip()
        if not raw:
            continue
        word = _clean_word(raw)
        if word:
            words.add(word)
        elif len(invalid) < 8:
            invalid.append(raw)
    return words, invalid


def parse_wordlist_content_ordered(content: str) -> list[str]:
    words = []
    seen = set()
    for token in re.split(r"[\s,;]+", content):
        word = _clean_word(token)
        if word and word not in seen:
            seen.add(word)
            words.append(word)
    return words


def _regular_inflection_candidates(word: str) -> set[str]:
    candidates = {word, f"{word}s", f"{word}ed", f"{word}ing", f"{word}er", f"{word}est"}
    if word.endswith("e"):
        candidates.update({f"{word}d", f"{word[:-1]}ing", f"{word}r", f"{word}st"})
    if len(word) > 2 and word.endswith("y") and word[-2] not in "aeiou":
        candidates.update({f"{word[:-1]}ies", f"{word[:-1]}ied", f"{word[:-1]}ier", f"{word[:-1]}iest"})
    if word.endswith(("s", "x", "z", "ch", "sh", "o")):
        candidates.add(f"{word}es")
    if word.endswith("ie"):
        candidates.add(f"{word[:-2]}ying")
    if (
        len(word) >= 3
        and word[-1] not in "aeiouwxy"
        and word[-2] in "aeiou"
        and word[-3] not in "aeiou"
    ):
        doubled = f"{word}{word[-1]}"
        candidates.update({f"{doubled}ed", f"{doubled}ing", f"{doubled}er", f"{doubled}est"})
    candidates.update(_IRREGULAR_INFLECTIONS.get(word, set()))
    return candidates


@lru_cache(maxsize=1)
def load_local_word_families() -> dict[str, tuple[str, ...]]:
    """Map BNC/COCA headwords and attested forms to their local word family."""
    if not BNC_COCA_PATH.exists():
        return {}
    families: dict[str, set[str]] = {}
    with BNC_COCA_PATH.open(encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            headword = _clean_word(str(row.get("headword") or ""))
            if not headword:
                continue
            forms = {headword}
            for part in str(row.get("word_forms") or "").split("),"):
                form = _clean_word(part.strip().split("(", 1)[0])
                if form:
                    forms.add(form)
            inflections = forms & _regular_inflection_candidates(headword)
            inflections.add(headword)
            for form in inflections:
                families.setdefault(form, set()).update(inflections)
    return {word: tuple(sorted(forms)) for word, forms in families.items()}


def expand_with_local_word_families(words: list[str]) -> dict[str, list[str]]:
    families = load_local_word_families()
    return {
        word: list(families[word])
        for word in words
        if word in families and len(families[word]) > 1
    }

def validate_wordlist_upload(filename: str, content: str):
    suffix = Path(filename).suffix.lower()
    if suffix not in {".txt", ".csv"}:
        return False, "词表只支持 .txt 或 .csv 文件。", set(), []
    words, invalid = parse_wordlist_content(content)
    if not words:
        return False, "没有识别到有效英文单词。请使用换行、空格或逗号分隔，例如：analyze, analyzed, analyzing。", words, invalid
    return True, "", words, invalid

def _split_triggers(value):
    if isinstance(value, list):
        return [str(t).strip() for t in value if str(t).strip()]
    return [t.strip() for t in re.split(r"[\n;|]+", str(value or "")) if t.strip()]

def _normalize_pattern_rows(rows):
    patterns = []
    for row in rows:
        item = {field: str(row.get(field, "") or "").strip() for field in REQUIRED_PATTERN_FIELDS if field != "triggers"}
        item["triggers"] = _split_triggers(row.get("triggers", ""))
        patterns.append(item)
    return patterns

def parse_patterns_upload(filename: str, raw: bytes):
    suffix = Path(filename).suffix.lower()
    if suffix not in PATTERN_UPLOAD_SUFFIXES:
        return False, "句式表支持 .csv、.xlsx 或 .json 文件。", []

    try:
        if suffix == ".json":
            data = json.loads(raw.decode("utf-8-sig", errors="replace"))
        elif suffix == ".csv":
            text = raw.decode("utf-8-sig", errors="replace")
            reader = csv.DictReader(io.StringIO(text))
            if not reader.fieldnames:
                return False, "CSV 第一行必须是表头。", []
            data = _normalize_pattern_rows(reader)
        else:
            try:
                from openpyxl import load_workbook
            except ImportError:
                return False, "读取 .xlsx 需要安装 openpyxl。", []
            wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            headers = [str(v or "").strip() for v in next(rows_iter, [])]
            if not any(headers):
                return False, "Excel 第一行必须是表头。", []
            rows = []
            for values in rows_iter:
                if not values or not any(v is not None and str(v).strip() for v in values):
                    continue
                rows.append({headers[i]: values[i] if i < len(values) else "" for i in range(len(headers))})
            data = _normalize_pattern_rows(rows)
    except json.JSONDecodeError as e:
        return False, f"JSON 解析失败：第 {e.lineno} 行第 {e.colno} 列，{e.msg}。", []
    except Exception as e:
        return False, f"句式表读取失败：{e}", []

    return validate_patterns_upload(data)

def validate_patterns_upload(data):
    """Validate normalized sentence pattern objects."""
    if isinstance(data, str):
        try:
            data = json.loads(data)
        except json.JSONDecodeError as e:
            return False, f"JSON 解析失败：第 {e.lineno} 行第 {e.colno} 列，{e.msg}。", []

    if not isinstance(data, list):
        return False, "句式文件必须能解析为多行记录。", []
    if not data:
        return False, "句式数组不能为空。", []

    seen_ids = set()
    for idx, item in enumerate(data, 1):
        if not isinstance(item, dict):
            return False, f"第 {idx} 项必须是对象。", []
        missing = [field for field in REQUIRED_PATTERN_FIELDS if field not in item]
        if missing:
            return False, f"第 {idx} 项缺少字段：{', '.join(missing)}。", []
        if not str(item.get("id", "")).strip():
            return False, f"第 {idx} 项 id 不能为空。", []
        if item["id"] in seen_ids:
            return False, f"句式 id 重复：{item['id']}。", []
        seen_ids.add(item["id"])
        triggers = item.get("triggers")
        if not isinstance(triggers, list) or not triggers:
            return False, f"第 {idx} 项 triggers 必须是非空数组。", []
        if not all(isinstance(t, str) and t.strip() for t in triggers):
            return False, f"第 {idx} 项 triggers 只能包含非空字符串。", []
        for field in ("name", "template", "explanation_cn", "example", "ielts_tip"):
            if not isinstance(item.get(field), str) or not item[field].strip():
                return False, f"第 {idx} 项 {field} 必须是非空字符串。", []
    return True, "", data

def _safe_stem(stem: str) -> str:
    return re.sub(r"[^a-z0-9_]", "_", stem.lower()).strip("_") or "custom"

def _user_compiled_path(source: Path) -> Path:
    return COMPILED_DIR / f"user_{_safe_stem(source.stem)}.json"

def _pattern_meta_path(path: Path) -> Path:
    return path.with_name(path.name + ".meta")

def _clean_label(text: str, fallback: str, max_len: int = 12) -> str:
    value = re.sub(r"\s+", " ", str(text or "").strip())
    return (value or fallback)[:max_len]

def compile_user_wordlist(file_path: Path, display_name: str | None = None, tag: str | None = None):
    """将用户上传的 CSV 或文本文件编译为 compiled 目录下的 JSON。"""
    stem = file_path.stem
    words = set()
    
    # 简单的分词逻辑，支持逗号、换行等分隔符
    try:
        content = file_path.read_text(encoding="utf-8", errors="replace")
        words, _invalid = parse_wordlist_content(content)
        
        safe_stem = _safe_stem(stem)
        default_name = stem.replace("_", " ").replace("-", " ").title()
        display_name = _clean_label(display_name, default_name, 24)
        tag = _clean_label(tag, display_name, 10)
        target_json = _user_compiled_path(file_path)
        config = {
            "metadata": {
                "name": display_name,
                "type": "domain",
                "key": f"user_{safe_stem}",
                "color": "domain-user",
                "tag": tag,
                "tag_class": "user-tag",
                "source": file_path.name,
            },
            "words": sorted(list(words))
        }
        
        with open(target_json, "w", encoding="utf-8") as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        print(f"Error compiling wordlist {file_path}: {e}")
        return False

def _safe_existing_child(directory: Path, filename: str) -> Path | None:
    if not filename or Path(filename).name != filename:
        return None
    if not re.fullmatch(r"[A-Za-z0-9_\-\.]+", filename):
        return None
    path = (directory / filename).resolve()
    try:
        path.relative_to(directory.resolve())
    except ValueError:
        return None
    return path if path.exists() else None

def list_uploaded_resources():
    wordlists = []
    for source in sorted(USER_DIR.glob("*")):
        if not source.is_file() or source.name == ".gitkeep":
            continue
        compiled = _user_compiled_path(source)
        count = 0
        metadata = {}
        if compiled.exists():
            try:
                data = json.loads(compiled.read_text(encoding="utf-8"))
                count = len(data.get("words", []))
                metadata = data.get("metadata", {}) if isinstance(data, dict) else {}
            except Exception:
                count = 0
        wordlists.append({
            "filename": source.name,
            "compiled": compiled.name if compiled.exists() else "",
            "key": metadata.get("key") or (compiled.stem if compiled.exists() else ""),
            "count": count,
            "tag": (metadata.get("name", source.stem)[:10] if metadata.get("tag") in {"", "USER", None} else metadata.get("tag")),
            "name": metadata.get("name", source.stem),
            "size": source.stat().st_size,
        })

    patterns = []
    for path in sorted(PATTERNS_DIR.glob("*.json")):
        if not path.is_file() or path.name == ".gitkeep":
            continue
        count = 0
        name = path.stem
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
            count = len(data) if isinstance(data, list) else 0
        except Exception:
            count = 0
        meta_path = _pattern_meta_path(path)
        if meta_path.exists():
            try:
                metadata = json.loads(meta_path.read_text(encoding="utf-8"))
                if isinstance(metadata, dict):
                    name = metadata.get("name") or name
            except Exception:
                pass
        patterns.append({
            "filename": path.name,
            "name": name,
            "count": count,
            "size": path.stat().st_size,
        })
    return {"wordlists": wordlists, "patterns": patterns}

def update_uploaded_wordlist_metadata(filename: str, name: str | None = None, tag: str | None = None):
    source = _safe_existing_child(USER_DIR, filename)
    if not source:
        return False, "wordlist not found", {}

    compiled = _user_compiled_path(source)
    if not compiled.exists():
        return False, "compiled wordlist not found", {}

    try:
        data = json.loads(compiled.read_text(encoding="utf-8"))
    except Exception:
        return False, "compiled wordlist metadata is not readable", {}

    if not isinstance(data, dict) or not isinstance(data.get("words"), list):
        return False, "compiled wordlist format is invalid", {}

    metadata = data.get("metadata")
    if not isinstance(metadata, dict):
        metadata = {}
        data["metadata"] = metadata

    current_name = metadata.get("name") or source.stem
    new_name = _clean_label(name, current_name, 24)
    new_tag = _clean_label(tag, metadata.get("tag") or new_name, 10)
    metadata["name"] = new_name
    metadata["tag"] = new_tag

    try:
        compiled.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception:
        return False, "failed to update compiled wordlist metadata", {}

    return True, "", {
        "filename": source.name,
        "name": new_name,
        "tag": new_tag,
        "count": len(data.get("words", [])),
    }

def update_uploaded_pattern_metadata(filename: str, name: str | None = None):
    path = _safe_existing_child(PATTERNS_DIR, filename)
    if not path:
        return False, "pattern file not found", {}
    if path.suffix.lower() != ".json":
        return False, "only uploaded pattern JSON files can be edited", {}

    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return False, "pattern file is not readable", {}
    if not isinstance(data, list):
        return False, "pattern file format is invalid", {}

    new_name = _clean_label(name, path.stem, 24)
    try:
        _pattern_meta_path(path).write_text(
            json.dumps({"name": new_name}, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    except Exception:
        return False, "failed to update pattern metadata", {}

    return True, "", {
        "filename": path.name,
        "name": new_name,
        "count": len(data),
    }

def delete_uploaded_wordlist(filename: str):
    source = _safe_existing_child(USER_DIR, filename)
    if not source:
        return False, "词表文件不存在。"
    compiled = _user_compiled_path(source)
    source.unlink()
    compiled.unlink(missing_ok=True)
    return True, ""

def delete_uploaded_pattern(filename: str):
    path = _safe_existing_child(PATTERNS_DIR, filename)
    if not path:
        return False, "句式表文件不存在。"
    if path.suffix.lower() != ".json":
        return False, "只能删除已上传的句式 JSON 文件。"
    path.unlink()
    _pattern_meta_path(path).unlink(missing_ok=True)
    return True, ""

# 默认配置，如果文件没有元数据则使用这些
DEFAULT_CONFIG = {
    "exclude_a1a2": {"name": "排除常用词", "type": "exclude", "key": "exclude"},
    "cefr_b1": {"name": "中级 (B1)", "type": "level", "key": "b1", "color": "level-b1", "tag": "B1"},
    "cefr_b2": {"name": "中高级 (B2)", "type": "level", "key": "b2", "color": "level-b2", "tag": "B2"},
    "cefr_c1": {"name": "高级 (C1)", "type": "level", "key": "c1", "color": "level-c1", "tag": "C1"},
    "domain_business": {"name": "商务", "type": "domain", "key": "business", "color": "domain-business", "tag": "商务"},
    "domain_academic": {"name": "学术", "type": "domain", "key": "academic", "color": "domain-academic", "tag": "学术"},
    "domain_fitness": {"name": "健身", "type": "domain", "key": "fitness", "color": "domain-fitness", "tag": "健身"},
    "domain_medical": {"name": "医学", "type": "domain", "key": "medical", "color": "domain-medical", "tag": "医学"},
}

def scan_wordlists():
    """扫描 compiled 目录下的所有词表 JSON 文件并提取元数据。"""
    if not COMPILED_DIR.exists():
        COMPILED_DIR.mkdir(parents=True, exist_ok=True)
    
    results = []
    # 优先扫描 compiled 目录
    for f in COMPILED_DIR.glob("*.json"):
        if f.name == "sentence_patterns.json":
            continue
        
        stem = f.stem
        config = DEFAULT_CONFIG.get(stem, {
            "name": stem.replace("_", " ").title(),
            "type": "domain" if "domain" in stem else "level",
            "key": stem.split("_")[-1],
            "color": f"domain-{stem.split('_')[-1]}",
            "tag": stem.split("_")[-1].upper()
        })
        
        try:
            with open(f, "r", encoding="utf-8") as jf:
                data = json.load(jf)
                # 如果文件自带 metadata 则覆盖默认
                if "metadata" in data:
                    config.update(data["metadata"])
                if stem.startswith("user_"):
                    display_name = config.get("name") or stem[5:].replace("_", " ").title()
                    config.update({
                        "name": display_name,
                        "type": "domain",
                        "key": stem,
                        "color": config.get("color") or "domain-user",
                        "tag": display_name[:10] if config.get("tag") in {"", "USER", None} else config.get("tag"),
                        "tag_class": config.get("tag_class") or "user-tag",
                    })
                
                config["id"] = stem
                config["count"] = len(data.get("words", []))
                results.append(config)
        except Exception:
            continue
            
    return sorted(results, key=lambda x: (x["type"] != "exclude", x["type"] != "level", x.get("id", "")))

def get_combined_patterns():
    """合并 patterns 目录下的所有句式 JSON 文件。"""
    if not PATTERNS_DIR.exists():
        PATTERNS_DIR.mkdir(parents=True, exist_ok=True)
    
    # 如果 legacy 文件存在，也包含进去
    legacy_file = COMPILED_DIR / "sentence_patterns.json"
    patterns = []
    
    if legacy_file.exists():
        try:
            with open(legacy_file, "r", encoding="utf-8") as f:
                data = json.load(f)
                if isinstance(data, list):
                    patterns.extend(data)
        except Exception:
            pass
            
    for f in PATTERNS_DIR.glob("*.json"):
        try:
            with open(f, "r", encoding="utf-8") as jf:
                data = json.load(jf)
                if isinstance(data, list):
                    patterns.extend(data)
        except Exception:
            continue
            
    return patterns
